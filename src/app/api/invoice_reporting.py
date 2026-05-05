"""
Invoice Reporting API – VAT invoice list (listado fiscal de facturas).

Security flow (same as /finance/ledger):
1. Validate JWT → get user_id (sub claim)
2. Fetch customer_profile from Supabase → get cta_contable
3. Query MARIADB_FINAN_DB via service/repository (NEVER expose cta_contable to client)

Note: This router is intentionally separate from api/invoices.py to avoid
modifying the existing /invoices endpoint.
"""
import asyncio
import logging
from datetime import date
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..core.auth import get_current_user, User
from ..core.date_utils import validate_date_range
from ..core.supabase_admin import fetch_customer_profile, SupabaseUnavailableError
from ..services.invoice_reporting_service import get_vat_invoice_list

logger = logging.getLogger(__name__)

_LOGO_PATH = Path(__file__).parent.parent / "assets" / "images" / "logo_santiago_vargas.png"

router = APIRouter(tags=["invoices"])


# ── Schemas ────────────────────────────────────────────────────

class VatInvoiceItem(BaseModel):
    fecha_fra: date
    num_fra: str
    base_imp: float
    tipo_iva: float
    cuota_iva: float
    tipo_recargo: float
    cuota_recargo: float
    imp_total: float


class VatInvoiceTotals(BaseModel):
    total_base: float
    total_iva: float
    total_recargo: float
    total_factura: float


class VatInvoiceListResponse(BaseModel):
    items: list[VatInvoiceItem]
    totals: VatInvoiceTotals


# ── Endpoint ───────────────────────────────────────────────────

@router.get("/invoices/vat-list", response_model=VatInvoiceListResponse)
async def get_vat_invoice_list_endpoint(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
) -> VatInvoiceListResponse:
    """
    Get the VAT invoice list (listado fiscal) for the authenticated customer.

    - Requires valid Supabase JWT
    - Both start_date and end_date are required
    - Client cannot specify cta_contable – resolved server-side from customer_profiles
    - Includes aggregated period totals (base, IVA, recargo, total)
    """
    try:
        validate_date_range(start_date, end_date)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    user_id = current_user.sub

    try:
        profile = await fetch_customer_profile(user_id)
    except SupabaseUnavailableError:
        raise HTTPException(status_code=503, detail="Upstream auth/profile service unavailable")

    if profile is None:
        logger.warning("No customer_profile found for user_id=%s", user_id)
        raise HTTPException(status_code=403, detail="No customer profile found")

    if not profile.is_active:
        logger.warning("Customer profile is_active=false for user_id=%s", user_id)
        raise HTTPException(status_code=403, detail="Customer profile is not active")

    if not profile.cta_contable:
        logger.warning("Customer profile has no cta_contable for user_id=%s", user_id)
        raise HTTPException(status_code=403, detail="Customer profile has no accounting account configured")

    try:
        result = await get_vat_invoice_list(
            cta_contable=profile.cta_contable,
            start_date=start_date,
            end_date=end_date,
        )
    except Exception as e:
        logger.error("Database error fetching VAT invoice list: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="Error fetching VAT invoice list")

    return result


# ── Excel helper ───────────────────────────────────────────────

_TOTALS_LABELS = [
    ("total_base", "Total Base"),
    ("total_iva", "Total IVA"),
    ("total_recargo", "Total Recargo"),
    ("total_factura", "Total Factura"),
]


def _build_xlsx(data: dict) -> BytesIO:
    # ── Style constants ──────────────────────────────────────────
    _THIN = Side(style="thin")
    _FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _TOP_BORDER = Border(top=_THIN)

    _HDR_FILL = PatternFill("solid", fgColor="F28C28")
    _HDR_FONT = Font(bold=True, color="FFFFFF")
    _BOLD = Font(bold=True)

    _CENTER = Alignment(horizontal="center", vertical="center")
    _RIGHT = Alignment(horizontal="right", vertical="center")

    FMT_CURRENCY = '#,##0.00 €'
    FMT_PERCENT = '0.00" %"'
    FMT_DATE = 'DD-MM-YYYY'

    # (visible header, field name, col width, number format, alignment)
    COLUMNS = [
        ("Fecha",          "fecha_fra",     14, FMT_DATE,     _CENTER),
        ("Nº Fra.",        "num_fra",        14, None,         _CENTER),
        ("Base Imponible", "base_imp",       18, FMT_CURRENCY, _RIGHT),
        ("% IVA",          "tipo_iva",       10, FMT_PERCENT,  _CENTER),
        ("Importe IVA",    "cuota_iva",      16, FMT_CURRENCY, _RIGHT),
        ("R.E.",           "tipo_recargo",   10, FMT_PERCENT,  _CENTER),
        ("Importe RE",     "cuota_recargo",  16, FMT_CURRENCY, _RIGHT),
        ("Total",          "imp_total",      16, FMT_CURRENCY, _RIGHT),
    ]
    _NUMERIC = {"base_imp", "tipo_iva", "cuota_iva", "tipo_recargo", "cuota_recargo", "imp_total"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # ── Header row ───────────────────────────────────────────────
    ws.append([col[0] for col in COLUMNS])
    ws.row_dimensions[1].height = 22
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER
        cell.border = _FULL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMNS[col_idx - 1][2]

    # ── Data rows ────────────────────────────────────────────────
    for item in data["items"]:
        row_idx = ws.max_row + 1
        ws.append([
            float(item[col[1]]) if col[1] in _NUMERIC else item[col[1]]
            for col in COLUMNS
        ])
        for col_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col[3]:
                cell.number_format = col[3]
            cell.alignment = col[4]

    # ── Autofilter + freeze pane ─────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    ws.freeze_panes = "A2"

    # ── Totals ───────────────────────────────────────────────────
    ws.append([])  # blank separator row

    totals = data["totals"]
    for i, (key, label) in enumerate(_TOTALS_LABELS):
        ws.append([label, float(totals[key])])
        row_idx = ws.max_row
        lc = ws.cell(row=row_idx, column=1)
        vc = ws.cell(row=row_idx, column=2)
        lc.font = _BOLD
        lc.alignment = _RIGHT
        vc.font = _BOLD
        vc.number_format = FMT_CURRENCY
        vc.alignment = _RIGHT
        if i == 0:
            lc.border = _TOP_BORDER
            vc.border = _TOP_BORDER

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Export endpoint ────────────────────────────────────────────

@router.get("/invoices/vat-list/export")
async def export_vat_invoice_list(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    try:
        validate_date_range(start_date, end_date)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    user_id = current_user.sub

    try:
        profile = await fetch_customer_profile(user_id)
    except SupabaseUnavailableError:
        raise HTTPException(status_code=503, detail="Upstream auth/profile service unavailable")

    if profile is None:
        raise HTTPException(status_code=403, detail="No customer profile found")

    if not profile.is_active:
        raise HTTPException(status_code=403, detail="Customer profile is not active")

    if not profile.cta_contable:
        raise HTTPException(status_code=403, detail="Customer profile has no accounting account configured")

    try:
        data = await get_vat_invoice_list(
            cta_contable=profile.cta_contable,
            start_date=start_date,
            end_date=end_date,
        )
    except Exception as e:
        logger.error("Database error fetching VAT invoice list for export: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="Error fetching VAT invoice list")

    buf = await asyncio.to_thread(_build_xlsx, data)

    filename = f"vat_invoices_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ── PDF helper ─────────────────────────────────────────────────

def _build_pdf(data: dict, start_date: date, end_date: date) -> BytesIO:
    _ORANGE      = rl_colors.HexColor("#F28C28")
    _ORANGE_SOFT = rl_colors.HexColor("#FEF0DB")
    _STRIPE      = rl_colors.HexColor("#FDF3E7")

    def _fmt_currency(v: float) -> str:
        negative = v < 0
        integer, decimal = f"{abs(v):,.2f}".split(".")
        result = f"{integer.replace(',', '.')},{decimal} €"
        return f"-{result}" if negative else result

    def _fmt_percent(v: float) -> str:
        return f"{v:.2f}".replace(".", ",") + " %"

    def _fmt_date(d) -> str:
        if hasattr(d, "strftime"):
            return d.strftime("%d-%m-%Y")
        if isinstance(d, str) and len(d) == 10:
            return f"{d[8:10]}-{d[5:7]}-{d[0:4]}"
        return str(d)

    # (visible header, field, width_mm, formatter)
    COLUMNS = [
        ("Fecha",          "fecha_fra",    24, _fmt_date),
        ("Nº Fra.",        "num_fra",      30, str),
        ("Base Imponible", "base_imp",     32, _fmt_currency),
        ("% IVA",          "tipo_iva",     18, _fmt_percent),
        ("Importe IVA",    "cuota_iva",    30, _fmt_currency),
        ("R.E.",           "tipo_recargo", 18, _fmt_percent),
        ("Importe RE",     "cuota_recargo",30, _fmt_currency),
        ("Total",          "imp_total",    30, _fmt_currency),
    ]
    col_widths  = [col[2] * mm for col in COLUMNS]
    table_total = sum(col_widths)          # 212 mm
    last_col_w  = col_widths[-1]           # 30 mm  (Total column)

    # ── Document styles ──────────────────────────────────────────
    base = getSampleStyleSheet()
    header_info_style = ParagraphStyle(
        "pdf_header_info",
        parent=base["Normal"],
        fontSize=9,
        fontName="Helvetica",
        leading=14,
    )
    client_style = ParagraphStyle(
        "pdf_client",
        parent=base["Normal"],
        fontSize=9,
        fontName="Helvetica-Bold",
        spaceAfter=4 * mm,
    )

    # ── Logo ─────────────────────────────────────────────────────
    img_r = ImageReader(str(_LOGO_PATH))
    img_px_w, img_px_h = img_r.getSize()
    logo_w = 55 * mm
    logo_h = logo_w * img_px_h / img_px_w
    logo_img = Image(str(_LOGO_PATH), width=logo_w, height=logo_h)

    # ── Header table: logo left | title + period right ───────────
    period_str = f"{start_date.strftime('%d-%m-%Y')} – {end_date.strftime('%d-%m-%Y')}"
    header_para = Paragraph(
        f'<font size="12"><b>Listado fiscal de facturas</b></font>'
        f'<br/><br/>Periodo: {period_str}',
        header_info_style,
    )
    logo_col_w = logo_w + 5 * mm
    header_table = Table(
        [[logo_img, header_para]],
        colWidths=[logo_col_w, table_total - logo_col_w],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Client block ─────────────────────────────────────────────
    cliente = next(
        (str(item.get("cliente", "")) for item in data["items"] if item.get("cliente")),
        "",
    )
    client_block = Table(
        [[Paragraph(f"Cliente: {cliente if cliente else '—'}", client_style)]],
        colWidths=[table_total],
    )
    client_block.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Main table ───────────────────────────────────────────────
    table_data = [[col[0] for col in COLUMNS]]
    for item in data["items"]:
        table_data.append([col[3](item[col[1]]) for col in COLUMNS])
    n_data = len(data["items"])

    ts_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0),      _ORANGE),
        ("TEXTCOLOR",  (0, 0), (-1, 0),      rl_colors.white),
        ("FONTNAME",   (0, 0), (-1, 0),      "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0),      8),
        ("ALIGN",      (0, 0), (-1, 0),      "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1),     "MIDDLE"),
        ("FONTNAME",   (0, 1), (-1, -1),     "Helvetica"),
        ("FONTSIZE",   (0, 1), (-1, -1),     8),
        ("GRID",       (0, 0), (-1, n_data), 0.3, rl_colors.grey),
        ("LINEBELOW",  (0, 0), (-1, 0),      0.7, rl_colors.black),
        ("ROWHEIGHT",  (0, 0), (-1, -1),     6 * mm),
        ("ALIGN", (0, 1), (1, n_data), "CENTER"),
        ("ALIGN", (2, 1), (2, n_data), "RIGHT"),
        ("ALIGN", (3, 1), (3, n_data), "CENTER"),
        ("ALIGN", (4, 1), (4, n_data), "RIGHT"),
        ("ALIGN", (5, 1), (5, n_data), "CENTER"),
        ("ALIGN", (6, 1), (7, n_data), "RIGHT"),
    ]
    for i in range(1, n_data + 1):
        if i % 2 == 0:
            ts_cmds.append(("BACKGROUND", (0, i), (-1, i), _STRIPE))

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle(ts_cmds))

    # ── Totals table — aligned with main table ───────────────────
    # label col = table_total - last_col_w so value col aligns with "Total" column
    totals = data["totals"]
    n_totals = len(_TOTALS_LABELS)
    totals_data = [
        [label, _fmt_currency(float(totals[key]))]
        for key, label in _TOTALS_LABELS
    ]
    totals_table = Table(
        totals_data,
        colWidths=[table_total - last_col_w, last_col_w],
    )
    totals_table.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (-1, -1),          "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1),          8),
        ("ALIGN",         (0, 0), (0, -1),           "RIGHT"),
        ("ALIGN",         (1, 0), (1, -1),           "RIGHT"),
        ("VALIGN",        (0, 0), (-1, -1),          "MIDDLE"),
        ("ROWHEIGHT",     (0, 0), (-1, -1),          6 * mm),
        ("TOPPADDING",    (0, 0), (-1, -1),          3),
        ("BOTTOMPADDING", (0, 0), (-1, -1),          3),
        ("RIGHTPADDING",  (0, 0), (-1, -1),          4),
        ("LINEABOVE",     (0, 0), (-1, 0),           0.7, rl_colors.black),
        # Highlight Total Factura (last row)
        ("BACKGROUND",    (0, n_totals - 1), (-1, n_totals - 1), _ORANGE_SOFT),
    ]))

    # ── Build document ───────────────────────────────────────────
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    doc.build([
        header_table,
        Spacer(1, 4 * mm),
        client_block,
        main_table,
        Spacer(1, 3 * mm),
        totals_table,
    ])
    buf.seek(0)
    return buf


# ── PDF export endpoint ────────────────────────────────────────

@router.get("/invoices/vat-list/export/pdf")
async def export_vat_invoice_list_pdf(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    try:
        validate_date_range(start_date, end_date)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    user_id = current_user.sub

    try:
        profile = await fetch_customer_profile(user_id)
    except SupabaseUnavailableError:
        raise HTTPException(status_code=503, detail="Upstream auth/profile service unavailable")

    if profile is None:
        raise HTTPException(status_code=403, detail="No customer profile found")

    if not profile.is_active:
        raise HTTPException(status_code=403, detail="Customer profile is not active")

    if not profile.cta_contable:
        raise HTTPException(status_code=403, detail="Customer profile has no accounting account configured")

    try:
        data = await get_vat_invoice_list(
            cta_contable=profile.cta_contable,
            start_date=start_date,
            end_date=end_date,
        )
    except Exception as e:
        logger.error("Database error fetching VAT invoice list for PDF export: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="Error fetching VAT invoice list")

    buf = await asyncio.to_thread(_build_pdf, data, start_date, end_date)

    filename = f"vat_invoices_{start_date}_{end_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
