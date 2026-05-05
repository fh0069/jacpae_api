"""
Finance API – Ledger/extracto del cliente.

Security flow (same as /invoices):
1. Validate JWT → get user_id (sub claim)
2. Fetch customer_profile from Supabase → get cta_contable
3. Query MARIADB_FINAN_DB via service/repository (NEVER expose cta_contable to client)
"""
import asyncio
import logging
from datetime import date
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..core.auth import get_current_user, User
from ..core.date_utils import validate_date_range
from ..core.supabase_admin import fetch_customer_profile, SupabaseUnavailableError
from ..services.finance_service import get_ledger

logger = logging.getLogger(__name__)

_LOGO_PATH = Path(__file__).parent.parent / "assets" / "images" / "logo_santiago_vargas.png"

router = APIRouter(tags=["finance"])


# ── Schemas ────────────────────────────────────────────────────

class LedgerItem(BaseModel):
    fecha: date
    concepto: str
    importe_debe: float
    importe_haber: float
    saldo: float


class LedgerResponse(BaseModel):
    start_date: date
    end_date: date
    exercise_start_date: date
    total_items: int
    items: list[LedgerItem]


# ── Endpoint ───────────────────────────────────────────────────

@router.get("/finance/ledger", response_model=LedgerResponse)
async def get_finance_ledger(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
) -> LedgerResponse:
    """
    Get the account ledger (extracto/mayor) for the authenticated customer.

    - Requires valid Supabase JWT
    - start_date and end_date define the visible range of movements
    - Opening balance is calculated from the start of start_date's fiscal year (service layer)
    - Client cannot specify cta_contable – resolved server-side from customer_profiles
    """
    try:
        validate_date_range(start_date, end_date)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    user_id = current_user.sub

    try:
        profile = await fetch_customer_profile(user_id)
    except SupabaseUnavailableError:
        raise HTTPException(status_code=503, detail="Upstream auth/profile service unavailable")

    if profile is None:
        logger.warning("No customer_profile found for user_id=%s", user_id)
        raise HTTPException(status_code=403, detail="No customer profile found")

    if not profile.is_active:
        logger.warning("Customer profile is_active=false for user_id=%s", user_id)
        raise HTTPException(status_code=403, detail="Customer profile is not active")

    if not profile.cta_contable:
        logger.warning("Customer profile has no cta_contable for user_id=%s", user_id)
        raise HTTPException(status_code=403, detail="Customer profile has no accounting account configured")

    try:
        result = await get_ledger(
            cta_contable=profile.cta_contable,
            start_date=start_date,
            end_date=end_date,
        )
    except Exception as e:
        logger.error("Database error fetching ledger: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="Error fetching ledger")

    return result


# ── Excel helper ───────────────────────────────────────────────

def _build_xlsx(data: dict) -> BytesIO:
    # ── Style constants ──────────────────────────────────────────
    _THIN = Side(style="thin")
    _FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    _HDR_FILL = PatternFill("solid", fgColor="F28C28")
    _HDR_FONT = Font(bold=True, color="FFFFFF")

    _CENTER = Alignment(horizontal="center", vertical="center")
    _LEFT   = Alignment(horizontal="left",   vertical="center")
    _RIGHT  = Alignment(horizontal="right",  vertical="center")

    FMT_CURRENCY = '#,##0.00 €'
    FMT_DATE     = 'DD-MM-YYYY'

    # (visible header, field, col width, number format, alignment)
    COLUMNS = [
        ("Fecha",    "fecha",         14, FMT_DATE,     _CENTER),
        ("Concepto", "concepto",      42, None,         _LEFT),
        ("Debe",     "importe_debe",  16, FMT_CURRENCY, _RIGHT),
        ("Haber",    "importe_haber", 16, FMT_CURRENCY, _RIGHT),
        ("Saldo",    "saldo",         16, FMT_CURRENCY, _RIGHT),
    ]
    _NUMERIC = {"importe_debe", "importe_haber", "saldo"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracto"

    # ── Header row ───────────────────────────────────────────────
    ws.append([col[0] for col in COLUMNS])
    ws.row_dimensions[1].height = 22
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER
        cell.border = _FULL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMNS[col_idx - 1][2]

    # ── Data rows ────────────────────────────────────────────────
    for item in data["items"]:
        row_idx = ws.max_row + 1
        ws.append([
            float(item[col[1]]) if col[1] in _NUMERIC else item[col[1]]
            for col in COLUMNS
        ])
        for col_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col[3]:
                cell.number_format = col[3]
            cell.alignment = col[4]

    # ── Autofilter + freeze pane ─────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Export endpoint ────────────────────────────────────────────

@router.get("/finance/ledger/export")
async def export_finance_ledger(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    try:
        validate_date_range(start_date, end_date)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    user_id = current_user.sub

    try:
        profile = await fetch_customer_profile(user_id)
    except SupabaseUnavailableError:
        raise HTTPException(status_code=503, detail="Upstream auth/profile service unavailable")

    if profile is None:
        raise HTTPException(status_code=403, detail="No customer profile found")

    if not profile.is_active:
        raise HTTPException(status_code=403, detail="Customer profile is not active")

    if not profile.cta_contable:
        raise HTTPException(status_code=403, detail="Customer profile has no accounting account configured")

    try:
        data = await get_ledger(
            cta_contable=profile.cta_contable,
            start_date=start_date,
            end_date=end_date,
        )
    except Exception as e:
        logger.error("Database error fetching ledger for export: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="Error fetching ledger")

    buf = await asyncio.to_thread(_build_xlsx, data)

    filename = f"ledger_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ── PDF helper ─────────────────────────────────────────────────

def _build_pdf(data: dict, start_date: date, end_date: date) -> BytesIO:
    _ORANGE = rl_colors.HexColor("#F28C28")
    _STRIPE = rl_colors.HexColor("#FDF3E7")

    def _fmt_currency(v: float) -> str:
        negative = v < 0
        integer, decimal = f"{abs(v):,.2f}".split(".")
        result = f"{integer.replace(',', '.')},{decimal} €"
        return f"-{result}" if negative else result

    def _fmt_date(d) -> str:
        if hasattr(d, "strftime"):
            return d.strftime("%d-%m-%Y")
        if isinstance(d, str) and len(d) == 10:
            return f"{d[8:10]}-{d[5:7]}-{d[0:4]}"
        return str(d)

    # col widths: Fecha 25 + Concepto 120 + Debe 35 + Haber 35 + Saldo 35 = 250 mm
    col_widths  = [25 * mm, 120 * mm, 35 * mm, 35 * mm, 35 * mm]
    table_total = sum(col_widths)  # 250 mm

    # ── Document styles ──────────────────────────────────────────
    base = getSampleStyleSheet()
    header_info_style = ParagraphStyle(
        "ledger_header_info",
        parent=base["Normal"],
        fontSize=9,
        fontName="Helvetica",
        leading=14,
    )
    client_style = ParagraphStyle(
        "ledger_client",
        parent=base["Normal"],
        fontSize=9,
        fontName="Helvetica-Bold",
        spaceAfter=4 * mm,
    )
    cell_wrap_style = ParagraphStyle(
        "ledger_cell",
        parent=base["Normal"],
        fontSize=8,
        fontName="Helvetica",
        leading=10,
    )
    summary_style = ParagraphStyle(
        "ledger_summary",
        parent=base["Normal"],
        fontSize=8,
        fontName="Helvetica",
        textColor=rl_colors.HexColor("#555555"),
    )

    # ── Logo ─────────────────────────────────────────────────────
    img_r = ImageReader(str(_LOGO_PATH))
    img_px_w, img_px_h = img_r.getSize()
    logo_w = 55 * mm
    logo_h = logo_w * img_px_h / img_px_w
    logo_img = Image(str(_LOGO_PATH), width=logo_w, height=logo_h)

    # ── Header table: logo left | title + period right ───────────
    period_str = f"{start_date.strftime('%d-%m-%Y')} – {end_date.strftime('%d-%m-%Y')}"
    header_para = Paragraph(
        f'<font size="12"><b>Extracto contable</b></font>'
        f'<br/><br/>Periodo: {period_str}',
        header_info_style,
    )
    logo_col_w = logo_w + 5 * mm
    header_table = Table(
        [[logo_img, header_para]],
        colWidths=[logo_col_w, table_total - logo_col_w],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Client block — same width as main table ───────────────────
    cliente = next(
        (item.get("cliente") for item in data["items"] if item.get("cliente")),
        "—",
    )
    client_block = Table(
        [[Paragraph(f"Cliente: {cliente}", client_style)]],
        colWidths=[table_total],
    )
    client_block.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Main table ───────────────────────────────────────────────
    table_data = [["Fecha", "Concepto", "Debe", "Haber", "Saldo"]]
    for item in data["items"]:
        table_data.append([
            _fmt_date(item["fecha"]),
            Paragraph(str(item["concepto"]), cell_wrap_style),
            _fmt_currency(float(item["importe_debe"])),
            _fmt_currency(float(item["importe_haber"])),
            _fmt_currency(float(item["saldo"])),
        ])
    n_data = len(data["items"])

    ts_cmds = [
        ("BACKGROUND",    (0, 0), (-1, 0),      _ORANGE),
        ("TEXTCOLOR",     (0, 0), (-1, 0),      rl_colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),      "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),      8),
        ("ALIGN",         (0, 0), (-1, 0),      "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1),     "MIDDLE"),
        ("FONTNAME",      (0, 1), (-1, -1),     "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1),     8),
        ("GRID",          (0, 0), (-1, n_data), 0.3, rl_colors.grey),
        ("LINEBELOW",     (0, 0), (-1, 0),      0.7, rl_colors.black),
        ("TOPPADDING",    (0, 0), (-1, -1),     3),
        ("BOTTOMPADDING", (0, 0), (-1, -1),     3),
        ("LEFTPADDING",   (0, 0), (-1, -1),     4),
        ("RIGHTPADDING",  (0, 0), (-1, -1),     4),
        ("ALIGN", (0, 1), (0, n_data), "CENTER"),  # Fecha
        ("ALIGN", (1, 1), (1, n_data), "LEFT"),    # Concepto
        ("ALIGN", (2, 1), (4, n_data), "RIGHT"),   # Debe, Haber, Saldo
    ]
    for i in range(1, n_data + 1):
        if i % 2 == 0:
            ts_cmds.append(("BACKGROUND", (0, i), (-1, i), _STRIPE))

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle(ts_cmds))

    # ── Summary — aligned with main table ────────────────────────
    summary_block = Table(
        [[Paragraph(f"Total movimientos: {data.get('total_items', n_data)}", summary_style)]],
        colWidths=[table_total],
    )
    summary_block.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Build document ───────────────────────────────────────────
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    doc.build([
        header_table,
        Spacer(1, 4 * mm),
        client_block,
        main_table,
        Spacer(1, 3 * mm),
        summary_block,
    ])
    buf.seek(0)
    return buf


# ── PDF export endpoint ────────────────────────────────────────

@router.get("/finance/ledger/export/pdf")
async def export_finance_ledger_pdf(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    try:
        validate_date_range(start_date, end_date)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    user_id = current_user.sub

    try:
        profile = await fetch_customer_profile(user_id)
    except SupabaseUnavailableError:
        raise HTTPException(status_code=503, detail="Upstream auth/profile service unavailable")

    if profile is None:
        raise HTTPException(status_code=403, detail="No customer profile found")

    if not profile.is_active:
        raise HTTPException(status_code=403, detail="Customer profile is not active")

    if not profile.cta_contable:
        raise HTTPException(status_code=403, detail="Customer profile has no accounting account configured")

    try:
        data = await get_ledger(
            cta_contable=profile.cta_contable,
            start_date=start_date,
            end_date=end_date,
        )
    except Exception as e:
        logger.error("Database error fetching ledger for PDF export: %s", type(e).__name__)
        raise HTTPException(status_code=500, detail="Error fetching ledger")

    buf = await asyncio.to_thread(_build_pdf, data, start_date, end_date)

    filename = f"ledger_{start_date}_{end_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
